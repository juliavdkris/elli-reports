import openpyxl
from docxtpl import DocxTemplate
from docx_charts import Document

import logging
from concurrent.futures import ProcessPoolExecutor


SPREADSHEET = 'original/PersRep_Data Pseud.xlsx'
TEMPLATE = 'original/template.docx'
OUTPUT_DIR = 'out'

logging.basicConfig(level=logging.INFO, format='%(message)s')

Entry = dict[str, str|float|None]


def parse_sheet_entries(filename: str) -> list[Entry]:
	wb = openpyxl.load_workbook(filename)
	sheet = wb[wb.sheetnames[0]]
	student_identification = wb['student_identification']

	entries: list[Entry] = []

	cols = [str(cell.value) for cell in sheet[1]] + [str(cell.value) for cell in student_identification[1]]
	rows = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True))
	sids = list(student_identification.iter_rows(min_row=2, max_row=student_identification.max_row, values_only=True))
	for i, (row, sid) in enumerate(zip(rows, sids)):
		entry: Entry = {}
		#! WARNING: this assumes that the columns in the main sheet are in the same order as in the studenet identification sheet
		for col, val in zip(cols, row+sid):
			if col is None:
				continue
			if val is None:
				entry[col] = None
				continue
			if entry.get(col) and entry[col] != str(val):
				logging.error(f'[!] Duplicate column name with different values not allowed: {col} on row {i+1} ({entry[col]} != {val})')
				exit(1)
			entry[col] = float(str(val)) if col.startswith('num_') else str(val)
		entries.append(entry)

	if len(entries) < len(rows):
		logging.warning(f'[!] No identification was found for {len(rows) - len(sids)}/{len(rows)} students')

	return entries


def to_valid_variable_name(s: str) -> str:
	s = s.replace('/', '_')  # Most of this script uses slashes as delimiter, but those aren't allowed in variable names
	s = ''.join([w if i == 0 else w.capitalize() for i, w in enumerate(s.split(' '))])  # Convert spaces to camelCase
	return s


def underscores_to_dict(strings: dict[str, float|None]) -> dict[str, dict[str, dict[str, float|None]]]:
	'''
	Example input: {'num/a/student/1': 1, 'num/a/student/2': 2, 'num/a/mean/1': 3, 'num/a/mean/2': 4, 'num/c/1': 10, 'num/c/2': 20, 'num/c/3': 30}
	Example output: {'a': {'student': {1: 1, 2: 2}, 'mean': {1: 3, 2: 4}}, 'c': {1: 10, 2: 20, 3: 30}
	'''

	# Terminology.:
	# 	- chart: the name of the metric, e.g. 'ce' for 'cognitive engagement'
	# 	- series: the row of entries, e.g. 'My score' or 'Total score'
	# 	- category: the column of entries, e.g. 'Week1', 'Week2'

	result: dict[str, dict[str, dict[str, float|None]]] = {}
	for key, value in strings.items():
		if not key.startswith('num/'):
			continue
		split = key.split('/')
		chart = split[1]
		series = split[2] if len(split) == 4 else 'Series1'
		category = split[-1]
		if chart not in result:
			result[chart] = {}
		if series not in result[chart]:
			result[chart][series] = {}
		result[chart][series][category] = value
	return result


def generate_report(entry: Entry, template: str, output_dir: str) -> None:
	# --> Stage 1: replace text in template
	filename = f'{output_dir}/{entry.get("new_id")}.docx'
	doc = DocxTemplate(template)
	context = {to_valid_variable_name(k): round(float(v), 2) if v and k.startswith('num/') else v for k, v in entry.items()}
	doc.render(context)
	doc.save(filename)
	logging.debug(f'    [*] Created string-replaced document for {entry.get("new_id")} ({entry.get("student_name")})')

	# --> Stage 2: update charts in generated report
	doc = Document(filename)

	# For each category of data (of an entry in the spreadsheet), update the corresponding chart in the generated report
	categories = underscores_to_dict({k: float(v) if isinstance(v, str) else None
										for k, v in entry.items() if k.startswith('num/')})
	for chart_name, value in categories.items():
		chart = doc.charts_by_name(chart_name)[0]
		data = chart.data()
		for series_name, new_series in value.items():
			for category, new_value in new_series.items():
				data[series_name][category] = new_value
				logging.debug(f'    [*] Updated chart data: {chart_name} ({series_name} -> {category} = {new_value})')
		chart.write(data)
	doc.save(filename)
	logging.debug(f'    [*] Updated charts for {entry.get("new_id")} ({entry.get("student_name")})')

	logging.info(f'[*] Generated report for {entry.get("new_id")} ({entry.get("student_name")})')


if __name__ == '__main__':
	with ProcessPoolExecutor() as executor:
		for entry in parse_sheet_entries(SPREADSHEET):
			executor.submit(generate_report, entry, TEMPLATE, OUTPUT_DIR)
